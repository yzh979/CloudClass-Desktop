#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
基础英语850词汇变形生成器
根据Ogden规则生成所有语法变形
"""

import json
import os
import csv
from pathlib import Path
from datetime import datetime
from jinja2 import Template
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 不规则动词变化表
IRREGULAR_VERBS = {
    'be': {'third': 'is', 'past': 'was/were', 'present_participle': 'being', 'past_participle': 'been'},
    'come': {'third': 'comes', 'past': 'came', 'present_participle': 'coming', 'past_participle': 'come'},
    'do': {'third': 'does', 'past': 'did', 'present_participle': 'doing', 'past_participle': 'done'},
    'get': {'third': 'gets', 'past': 'got', 'present_participle': 'getting', 'past_participle': 'got/gotten'},
    'give': {'third': 'gives', 'past': 'gave', 'present_participle': 'giving', 'past_participle': 'given'},
    'go': {'third': 'goes', 'past': 'went', 'present_participle': 'going', 'past_participle': 'gone'},
    'have': {'third': 'has', 'past': 'had', 'present_participle': 'having', 'past_participle': 'had'},
    'make': {'third': 'makes', 'past': 'made', 'present_participle': 'making', 'past_participle': 'made'},
    'say': {'third': 'says', 'past': 'said', 'present_participle': 'saying', 'past_participle': 'said'},
    'see': {'third': 'sees', 'past': 'saw', 'present_participle': 'seeing', 'past_participle': 'seen'},
    'send': {'third': 'sends', 'past': 'sent', 'present_participle': 'sending', 'past_participle': 'sent'},
    'take': {'third': 'takes', 'past': 'took', 'present_participle': 'taking', 'past_participle': 'taken'},
}

# 不规则名词复数
IRREGULAR_PLURALS = {
    'man': 'men',
    'woman': 'women',
    'child': 'children',
    'tooth': 'teeth',
    'foot': 'feet',
    'person': 'people',
    'leaf': 'leaves',
    'knife': 'knives',
    'sheep': 'sheep',
    'fish': 'fish',
}

# 不规则形容词比较级
IRREGULAR_COMPARISONS = {
    'good': {'comparative': 'better', 'superlative': 'best'},
    'bad': {'comparative': 'worse', 'superlative': 'worst'},
    'little': {'comparative': 'less', 'superlative': 'least'},
    'much': {'comparative': 'more', 'superlative': 'most'},
    'far': {'comparative': 'farther/further', 'superlative': 'farthest/furthest'},
}


def pluralize_noun(word):
    """
    规则1：生成名词复数
    
    规则：
    - 一般名词 + s: book → books
    - 以s/x/ch/sh结尾 + es: box → boxes
    - 辅音+y结尾变ies: baby → babies
    - 以f/fe结尾变ves: knife → knives
    - 不规则变化
    
    Args:
        word (str): 原词
        
    Returns:
        str: 复数形式
    """
    # 检查不规则复数
    if word in IRREGULAR_PLURALS:
        return IRREGULAR_PLURALS[word]
    
    # 以s, x, ch, sh结尾加es
    if word.endswith(('s', 'x', 'ch', 'sh', 'ss')):
        return word + 'es'
    
    # 辅音+y结尾变ies
    if len(word) > 1 and word.endswith('y') and word[-2] not in 'aeiou':
        return word[:-1] + 'ies'
    
    # 以f/fe结尾变ves
    if word.endswith('f'):
        return word[:-1] + 'ves'
    if word.endswith('fe'):
        return word[:-2] + 'ves'
    
    # 以o结尾（辅音+o加es，元音+o加s）
    if word.endswith('o') and len(word) > 1 and word[-2] not in 'aeiou':
        return word + 'es'
    
    # 一般情况加s
    return word + 's'


def derive_er_form(word):
    """
    规则2a：-er派生（做...的人/物）
    
    规则：
    - 一般词 + er: teach → teacher
    - 以e结尾 + r: make → maker
    - 辅音+y变ier: carry → carrier
    - 重读闭音节双写末尾字母+er: run → runner
    
    Args:
        word (str): 原词
        
    Returns:
        str: -er派生形式
    """
    # 以e结尾，直接加r
    if word.endswith('e'):
        return word + 'r'
    
    # 辅音+y结尾变ier
    if len(word) > 1 and word.endswith('y') and word[-2] not in 'aeiou':
        return word[:-1] + 'ier'
    
    # 重读闭音节（简化处理：单音节，元音+辅音结尾）
    if len(word) >= 3 and word[-1] not in 'aeiouywxh' and word[-2] in 'aeiou' and word[-3] not in 'aeiou':
        return word + word[-1] + 'er'
    
    # 一般情况加er
    return word + 'er'


def derive_ing_form(word):
    """
    规则2b：-ing派生（名词性/形容词性）
    
    规则：
    - 一般词 + ing: build → building
    - 以不发音的e结尾，去e加ing: make → making
    - 重读闭音节双写末尾字母+ing: run → running
    
    Args:
        word (str): 原词
        
    Returns:
        str: -ing派生形式
    """
    # 以ie结尾变ying
    if word.endswith('ie'):
        return word[:-2] + 'ying'
    
    # 以不发音的e结尾（非ee, oe等），去e加ing
    if word.endswith('e') and not word.endswith(('ee', 'oe', 'ye')):
        return word[:-1] + 'ing'
    
    # 重读闭音节双写
    if len(word) >= 3 and word[-1] not in 'aeiouywxh' and word[-2] in 'aeiou' and word[-3] not in 'aeiou':
        return word + word[-1] + 'ing'
    
    # 一般情况加ing
    return word + 'ing'


def derive_ed_form(word):
    """
    规则2c：-ed派生（形容词：被...的）
    
    规则：
    - 一般词 + ed: interest → interested
    - 以e结尾 + d: excite → excited
    - 辅音+y变ied: worry → worried
    - 重读闭音节双写末尾字母+ed: stop → stopped
    
    Args:
        word (str): 原词
        
    Returns:
        str: -ed派生形式
    """
    # 以e结尾，直接加d
    if word.endswith('e'):
        return word + 'd'
    
    # 辅音+y结尾变ied
    if len(word) > 1 and word.endswith('y') and word[-2] not in 'aeiou':
        return word[:-1] + 'ied'
    
    # 重读闭音节双写
    if len(word) >= 3 and word[-1] not in 'aeiouywxh' and word[-2] in 'aeiou' and word[-3] not in 'aeiou':
        return word + word[-1] + 'ed'
    
    # 一般情况加ed
    return word + 'ed'


def derive_ly_adverb(word):
    """
    规则3：形容词派生副词（-ly）
    
    规则：
    - 一般形容词 + ly: quick → quickly
    - 以y结尾变ily: happy → happily
    - 以le结尾去e加y: simple → simply
    - 以ic结尾加ally: automatic → automatically
    
    Args:
        word (str): 形容词
        
    Returns:
        str: 副词形式
    """
    # 以ic结尾加ally
    if word.endswith('ic'):
        return word + 'ally'
    
    # 以le结尾去e加y
    if word.endswith('le') and len(word) > 2:
        return word[:-1] + 'y'
    
    # 以y结尾变ily
    if word.endswith('y') and len(word) > 1 and word[-2] not in 'aeiou':
        return word[:-1] + 'ily'
    
    # 一般情况加ly
    return word + 'ly'


def get_comparative_superlative(word):
    """
    规则4：比较级和最高级
    
    规则：
    - 单音节和部分双音节：-er/-est (small → smaller → smallest)
    - 长词用more/most: beautiful → more beautiful → most beautiful
    - 不规则变化
    
    Args:
        word (str): 形容词
        
    Returns:
        tuple: (comparative, superlative)
    """
    # 检查不规则变化
    if word in IRREGULAR_COMPARISONS:
        return (IRREGULAR_COMPARISONS[word]['comparative'], 
                IRREGULAR_COMPARISONS[word]['superlative'])
    
    # 判断音节数（简化：根据长度判断）
    # 单音节或以y结尾的双音节使用-er/-est
    is_short = len(word) <= 6 or (word.endswith('y') and len(word) <= 7)
    
    if is_short:
        # 以e结尾加r/st
        if word.endswith('e'):
            comparative = word + 'r'
            superlative = word + 'st'
        # 辅音+y变ier/iest
        elif word.endswith('y') and len(word) > 1 and word[-2] not in 'aeiou':
            comparative = word[:-1] + 'ier'
            superlative = word[:-1] + 'iest'
        # 重读闭音节双写
        elif len(word) >= 3 and word[-1] not in 'aeiouywxh' and word[-2] in 'aeiou' and word[-3] not in 'aeiou':
            comparative = word + word[-1] + 'er'
            superlative = word + word[-1] + 'est'
        else:
            comparative = word + 'er'
            superlative = word + 'est'
    else:
        # 长词用more/most
        comparative = f'more {word}'
        superlative = f'most {word}'
    
    return (comparative, superlative)


def add_negative_un(word):
    """
    规则5：否定形容词（un-）
    
    规则：
    - 大部分形容词可加un-: happy → unhappy
    - 颜色形容词、比较级等不适用
    
    Args:
        word (str): 形容词
        
    Returns:
        str or None: 否定形式或None（不适用）
    """
    # 颜色形容词不加un-
    colors = ['black', 'white', 'red', 'blue', 'green', 'yellow', 'brown', 'gray', 'grey', 'purple', 'pink', 'orange']
    if word in colors:
        return None
    
    # 已经是否定的不再加
    if word.startswith('un'):
        return None
    
    # 某些形容词不适用un-（使用in-, im-, ir-, il-等）
    non_un_prefixes = ['possible', 'perfect', 'regular', 'responsible', 'legal', 'mature']
    if word in non_un_prefixes:
        return None
    
    # 比较级、最高级不加un-
    if word.endswith(('er', 'est')):
        return None
    
    return 'un' + word


def generate_compound_words(word, all_words):
    """
    规则6：生成复合词
    
    规则：
    - 名词+名词: milk + man → milkman
    - 名词+方向词: sun + down → sundown
    - 基于语义匹配生成常见复合词
    
    Args:
        word (str): 原词
        all_words (list): 所有词汇列表
        
    Returns:
        list: 复合词列表
    """
    compounds = []
    
    # 预定义的常见复合词规则
    compound_rules = {
        'sun': ['light', 'down', 'rise', 'set'],
        'moon': ['light'],
        'fire': ['man', 'place', 'work'],
        'water': ['way', 'fall', 'mark'],
        'book': ['mark', 'case', 'keeper'],
        'hand': ['writing', 'shake', 'work'],
        'work': ['man', 'shop', 'room'],
        'house': ['work', 'keeper', 'wife'],
        'foot': ['print', 'ball', 'step'],
        'head': ['light', 'line', 'way'],
        'back': ['bone', 'ground'],
        'day': ['light', 'time', 'break'],
        'night': ['time', 'fall'],
        'time': ['keeper'],
        'door': ['way', 'step'],
        'window': ['pane'],
    }
    
    if word in compound_rules:
        for suffix in compound_rules[word]:
            compounds.append(word + suffix)
    
    return compounds


def conjugate_verb(word):
    """
    规则7：动词变位
    
    生成：第三人称单数、过去式、现在分词、过去分词
    
    Args:
        word (str): 动词
        
    Returns:
        dict: 包含各种变位形式的字典
    """
    # 检查不规则动词
    if word in IRREGULAR_VERBS:
        return IRREGULAR_VERBS[word]
    
    result = {}
    
    # 第三人称单数
    if word.endswith(('s', 'x', 'ch', 'sh', 'o')):
        result['third'] = word + 'es'
    elif word.endswith('y') and len(word) > 1 and word[-2] not in 'aeiou':
        result['third'] = word[:-1] + 'ies'
    else:
        result['third'] = word + 's'
    
    # 过去式和过去分词（规则动词相同）
    if word.endswith('e'):
        past = word + 'd'
    elif word.endswith('y') and len(word) > 1 and word[-2] not in 'aeiou':
        past = word[:-1] + 'ied'
    elif len(word) >= 3 and word[-1] not in 'aeiouywxh' and word[-2] in 'aeiou' and word[-3] not in 'aeiou':
        past = word + word[-1] + 'ed'
    else:
        past = word + 'ed'
    
    result['past'] = past
    result['past_participle'] = past
    
    # 现在分词
    result['present_participle'] = derive_ing_form(word)
    
    return result


def process_word(word_data, all_words):
    """
    处理单个词汇，生成所有适用的变形
    
    Args:
        word_data (dict): 词汇数据
        all_words (list): 所有词汇列表
        
    Returns:
        dict: 包含原词和所有变形的字典
    """
    word = word_data['word']
    pos = word_data['pos']
    chinese = word_data['chinese']
    
    result = {
        'word': word,
        'pos': pos,
        'chinese': chinese,
        'plural': '',
        'comparative': '',
        'superlative': '',
        'er_form': '',
        'ing_form': '',
        'ed_form': '',
        'ly_form': '',
        'un_form': '',
        'compounds': [],
        'verb_forms': {},
        'rules_applied': []
    }
    
    # 名词处理
    if pos == 'noun':
        result['plural'] = pluralize_noun(word)
        result['rules_applied'].append('名词复数')
        
        # 300个名词的派生（标记为can_derive的）
        if word_data.get('can_derive', False):
            result['er_form'] = derive_er_form(word)
            result['ing_form'] = derive_ing_form(word)
            result['ed_form'] = derive_ed_form(word)
            result['rules_applied'].extend(['-er派生', '-ing派生', '-ed派生'])
        
        # 复合词
        compounds = generate_compound_words(word, all_words)
        if compounds:
            result['compounds'] = compounds
            result['rules_applied'].append('复合词')
    
    # 形容词处理
    if pos == 'adjective':
        # 副词派生
        result['ly_form'] = derive_ly_adverb(word)
        result['rules_applied'].append('副词派生-ly')
        
        # 比较级和最高级
        if word_data.get('can_compare', True):
            comparative, superlative = get_comparative_superlative(word)
            result['comparative'] = comparative
            result['superlative'] = superlative
            result['rules_applied'].append('比较级/最高级')
        
        # 否定形式
        un_form = add_negative_un(word)
        if un_form:
            result['un_form'] = un_form
            result['rules_applied'].append('否定un-')
    
    # 动词处理
    if pos == 'verb' and word_data.get('can_conjugate', False):
        verb_forms = conjugate_verb(word)
        result['verb_forms'] = verb_forms
        result['rules_applied'].append('动词变位')
    
    return result


def generate_html(data, template_path, output_path):
    """
    生成HTML交互页面
    
    Args:
        data (list): 处理后的词汇数据
        template_path (str): 模板文件路径
        output_path (str): 输出文件路径
    """
    with open(template_path, 'r', encoding='utf-8') as f:
        template_content = f.read()
    
    template = Template(template_content)
    
    # 统计信息
    stats = {
        'total_words': len(data),
        'nouns': sum(1 for item in data if item['pos'] == 'noun'),
        'adjectives': sum(1 for item in data if item['pos'] == 'adjective'),
        'verbs': sum(1 for item in data if item['pos'] == 'verb'),
        'with_plural': sum(1 for item in data if item['plural']),
        'with_er': sum(1 for item in data if item['er_form']),
        'with_ing': sum(1 for item in data if item['ing_form']),
        'with_ed': sum(1 for item in data if item['ed_form']),
        'with_ly': sum(1 for item in data if item['ly_form']),
        'with_un': sum(1 for item in data if item['un_form']),
        'with_comparative': sum(1 for item in data if item['comparative']),
        'with_compounds': sum(1 for item in data if item['compounds']),
        'with_verb_forms': sum(1 for item in data if item['verb_forms']),
    }
    
    html_content = template.render(
        data=data,
        stats=stats,
        generation_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    )
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)


def generate_csv(data, output_path):
    """
    生成CSV文件
    
    Args:
        data (list): 处理后的词汇数据
        output_path (str): 输出文件路径
    """
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        fieldnames = [
            '原词', '词性', '中文释义', '复数形式', '比较级', '最高级',
            '-er派生', '-ing派生', '-ed派生', '-ly派生', '否定形式',
            '复合词', '第三人称单数', '过去式', '现在分词', '过去分词',
            '变形规则说明'
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        
        for item in data:
            verb_forms = item.get('verb_forms', {})
            writer.writerow({
                '原词': item['word'],
                '词性': item['pos'],
                '中文释义': item['chinese'],
                '复数形式': item['plural'],
                '比较级': item['comparative'],
                '最高级': item['superlative'],
                '-er派生': item['er_form'],
                '-ing派生': item['ing_form'],
                '-ed派生': item['ed_form'],
                '-ly派生': item['ly_form'],
                '否定形式': item['un_form'],
                '复合词': '; '.join(item['compounds']),
                '第三人称单数': verb_forms.get('third', ''),
                '过去式': verb_forms.get('past', ''),
                '现在分词': verb_forms.get('present_participle', ''),
                '过去分词': verb_forms.get('past_participle', ''),
                '变形规则说明': '; '.join(item['rules_applied'])
            })


def generate_excel(data, output_path):
    """
    生成Excel文件（多个工作表）
    
    Args:
        data (list): 处理后的词汇数据
        output_path (str): 输出文件路径
    """
    wb = Workbook()
    
    # 删除默认工作表
    wb.remove(wb.active)
    
    # 样式定义
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Sheet 1: 完整派生词列表
    ws1 = wb.create_sheet('完整派生词列表')
    headers = [
        '原词', '词性', '中文释义', '复数形式', '比较级', '最高级',
        '-er派生', '-ing派生', '-ed派生', '-ly派生', '否定形式',
        '复合词', '第三人称单数', '过去式', '现在分词', '过去分词',
        '变形规则说明'
    ]
    ws1.append(headers)
    
    # 设置表头样式
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    for item in data:
        verb_forms = item.get('verb_forms', {})
        ws1.append([
            item['word'],
            item['pos'],
            item['chinese'],
            item['plural'],
            item['comparative'],
            item['superlative'],
            item['er_form'],
            item['ing_form'],
            item['ed_form'],
            item['ly_form'],
            item['un_form'],
            '; '.join(item['compounds']),
            verb_forms.get('third', ''),
            verb_forms.get('past', ''),
            verb_forms.get('present_participle', ''),
            verb_forms.get('past_participle', ''),
            '; '.join(item['rules_applied'])
        ])
    
    # Sheet 2: 仅名词
    ws2 = wb.create_sheet('名词及其变形')
    noun_headers = ['原词', '中文释义', '复数形式', '-er派生', '-ing派生', '-ed派生', '复合词']
    ws2.append(noun_headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    for item in data:
        if item['pos'] == 'noun':
            ws2.append([
                item['word'],
                item['chinese'],
                item['plural'],
                item['er_form'],
                item['ing_form'],
                item['ed_form'],
                '; '.join(item['compounds'])
            ])
    
    # Sheet 3: 仅形容词
    ws3 = wb.create_sheet('形容词及其变形')
    adj_headers = ['原词', '中文释义', '比较级', '最高级', '-ly派生', '否定un-']
    ws3.append(adj_headers)
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    for item in data:
        if item['pos'] == 'adjective':
            ws3.append([
                item['word'],
                item['chinese'],
                item['comparative'],
                item['superlative'],
                item['ly_form'],
                item['un_form']
            ])
    
    # Sheet 4: 仅动词
    ws4 = wb.create_sheet('动词及其变形')
    verb_headers = ['原词', '中文释义', '第三人称单数', '过去式', '现在分词', '过去分词']
    ws4.append(verb_headers)
    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    for item in data:
        if item['pos'] == 'verb' and item['verb_forms']:
            verb_forms = item['verb_forms']
            ws4.append([
                item['word'],
                item['chinese'],
                verb_forms.get('third', ''),
                verb_forms.get('past', ''),
                verb_forms.get('present_participle', ''),
                verb_forms.get('past_participle', '')
            ])
    
    # Sheet 5: 统计汇总
    ws5 = wb.create_sheet('统计汇总')
    ws5.append(['统计项', '数量'])
    for cell in ws5[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    stats = [
        ('总词数', len(data)),
        ('名词数', sum(1 for item in data if item['pos'] == 'noun')),
        ('形容词数', sum(1 for item in data if item['pos'] == 'adjective')),
        ('动词数', sum(1 for item in data if item['pos'] == 'verb')),
        ('其他词性数', sum(1 for item in data if item['pos'] not in ['noun', 'adjective', 'verb'])),
        ('有复数形式', sum(1 for item in data if item['plural'])),
        ('有-er派生', sum(1 for item in data if item['er_form'])),
        ('有-ing派生', sum(1 for item in data if item['ing_form'])),
        ('有-ed派生', sum(1 for item in data if item['ed_form'])),
        ('有-ly派生', sum(1 for item in data if item['ly_form'])),
        ('有un-否定', sum(1 for item in data if item['un_form'])),
        ('有比较级', sum(1 for item in data if item['comparative'])),
        ('有复合词', sum(1 for item in data if item['compounds'])),
        ('有动词变位', sum(1 for item in data if item['verb_forms'])),
    ]
    
    for stat in stats:
        ws5.append(stat)
    
    # Sheet 6: Ogden规则说明
    ws6 = wb.create_sheet('Ogden规则说明')
    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 80
    
    ws6.append(['规则编号', '规则说明'])
    for cell in ws6[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    rules = [
        ('规则1', '名词复数：一般加s，以s/x/ch/sh结尾加es，辅音+y变ies，不规则变化'),
        ('规则2a', '-er派生（做...的人/物）：一般加er，以e结尾加r，辅音+y变ier'),
        ('规则2b', '-ing派生（名词性/形容词性）：一般加ing，不发音e去e加ing，重读闭音节双写'),
        ('规则2c', '-ed派生（形容词：被...的）：一般加ed，以e结尾加d，辅音+y变ied'),
        ('规则3', '形容词派生副词-ly：一般加ly，以y结尾变ily，以le结尾去e加y，以ic结尾加ally'),
        ('规则4', '比较级和最高级：短词用-er/-est，长词用more/most，不规则变化'),
        ('规则5', '否定形容词un-：大部分形容词可加un-，颜色形容词除外'),
        ('规则6', '复合词：名词+名词，名词+方向词等常见组合'),
        ('规则7', '动词变位：第三人称单数、过去式、现在分词、过去分词，包括不规则动词'),
    ]
    
    for rule in rules:
        ws6.append(rule)
    
    # 调整列宽
    for ws in [ws1, ws2, ws3, ws4]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)


def main():
    """
    主函数：加载数据，处理所有词汇，生成输出文件
    """
    print("=" * 60)
    print("基础英语850词汇变形生成器")
    print("=" * 60)
    
    # 确定路径
    script_dir = Path(__file__).parent
    data_file = script_dir / 'basic_english_850.json'
    template_file = script_dir / 'templates' / 'index.html.j2'
    output_dir = script_dir / 'output'
    
    # 创建输出目录
    output_dir.mkdir(exist_ok=True)
    
    print(f"\n1. 加载词汇数据: {data_file}")
    
    # 加载数据
    with open(data_file, 'r', encoding='utf-8') as f:
        raw_data = json.load(f)
    
    # 合并所有词汇
    all_words_data = (
        raw_data['operations'] +
        raw_data['things_general'] +
        raw_data['things_picturable'] +
        raw_data['qualities']
    )
    
    all_words = [item['word'] for item in all_words_data]
    
    print(f"   加载了 {len(all_words_data)} 个词汇")
    print(f"   - Operations: {len(raw_data['operations'])}")
    print(f"   - General Things: {len(raw_data['things_general'])}")
    print(f"   - Picturable Things: {len(raw_data['things_picturable'])}")
    print(f"   - Qualities: {len(raw_data['qualities'])}")
    
    print("\n2. 处理词汇并生成变形...")
    
    # 处理所有词汇
    processed_data = []
    for word_data in all_words_data:
        result = process_word(word_data, all_words)
        processed_data.append(result)
    
    print(f"   处理完成，共生成 {len(processed_data)} 个词条")
    
    # 生成HTML
    print(f"\n3. 生成HTML交互页面: {output_dir / 'basic_english_derivatives.html'}")
    generate_html(processed_data, template_file, output_dir / 'basic_english_derivatives.html')
    print("   ✓ HTML页面生成成功")
    
    # 生成CSV
    print(f"\n4. 生成CSV文件: {output_dir / 'basic_english_derivatives.csv'}")
    generate_csv(processed_data, output_dir / 'basic_english_derivatives.csv')
    print("   ✓ CSV文件生成成功")
    
    # 生成Excel
    print(f"\n5. 生成Excel文件: {output_dir / 'basic_english_derivatives.xlsx'}")
    generate_excel(processed_data, output_dir / 'basic_english_derivatives.xlsx')
    print("   ✓ Excel文件生成成功")
    
    print("\n" + "=" * 60)
    print("所有文件生成完成！")
    print("=" * 60)
    print(f"\n输出目录: {output_dir}")
    print(f"  - basic_english_derivatives.html  (交互式HTML页面)")
    print(f"  - basic_english_derivatives.csv   (CSV数据文件)")
    print(f"  - basic_english_derivatives.xlsx  (Excel多表格文件)")
    print("\n请在浏览器中打开HTML文件查看交互界面。")
    print()


if __name__ == '__main__':
    main()
